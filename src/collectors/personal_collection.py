#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Scraper MOTU Origins -> Excel multi-hoja con hipervínculos a imágenes locales.
Ejecución directa en PyCharm o terminal.

- Rutas relativas al directorio del script:
    ./MOTU/              (Excel y datos)
    ./MOTU/images/       (descarga de imágenes)
- HTTP robusto: User-Agent realista (override por env MOTU_UA), reintentos, timeouts, pausas.
- Scraping tolerante a cambios menores.
- Mantiene el estado "Adquirido" desde un Excel previo si existe.
- Valida y colorea "Sí"/"No" en Excel, e inserta hipervínculos a imágenes locales.
- Nombres de hoja únicos (case-insensitive) con sufijos si hay colisiones.

Dependencias:
    pip install requests beautifulsoup4 pandas openpyxl xlsxwriter
"""

import os
import re
import time
import hashlib
import logging
import urllib.parse
from pathlib import Path
from typing import List, Tuple, Optional, Set

import requests
from requests.adapters import HTTPAdapter, Retry
from bs4 import BeautifulSoup
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook

# =========================
# Configuración y constantes
# =========================

CHECKLIST_URL = "https://www.actionfigure411.com/masters-of-the-universe/origins-checklist.php"
SITE_BASE = "https://www.actionfigure411.com/"
DESIRED_ORDER = [
    "Adquirido", "Name", "Wave", "Year", "Retail",
    "Imagen", "Image Path", "Image URL", "Detail Link"
]

REQUEST_TIMEOUT = 20.0     # segundos por petición
POLITE_SLEEP = 0.6         # pausa entre peticiones (cortesía)
MAX_RETRIES = 4            # reintentos para 429/5xx

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)

# =========================
# Utilidades de ruta
# =========================

def get_project_paths() -> Tuple[Path, Path, Path]:
    """
    Devuelve (project_root, excel_path, images_dir).
    project_root = carpeta de datos en data/MOTU
    """
    # Assuming src/collectors/personal_collection.py
    script_dir = Path(__file__).resolve().parent
    # Go up to src, then to project root
    project_base = script_dir.parent.parent
    
    project_root = project_base / "data" / "MOTU"
    images_dir = project_root / "images"
    excel_path = project_root / "lista_MOTU.xlsx"
    
    project_root.mkdir(parents=True, exist_ok=True)
    images_dir.mkdir(parents=True, exist_ok=True)
    return project_root, excel_path, images_dir

# =========================
# Sesión HTTP robusta
# =========================

def build_session() -> requests.Session:
    """
    Sesión HTTP robusta:
    - User-Agent realista tipo Chrome por defecto.
    - Permite override vía variable de entorno MOTU_UA.
    - Cabeceras de navegador comunes.
    - Reintentos con backoff para 429/5xx y respeto de Retry-After.
    """
    default_ua = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    user_agent = os.getenv("MOTU_UA", default_ua)

    s = requests.Session()
    s.headers.update({
        "User-Agent": user_agent,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
        "Connection": "keep-alive",
        "Referer": SITE_BASE,
    })

    retries = Retry(
        total=MAX_RETRIES,
        connect=3,
        read=3,
        backoff_factor=0.8,  # 0.8, 1.6, 3.2, 6.4 ...
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "HEAD", "OPTIONS"],
        raise_on_status=False,
        respect_retry_after_header=True,
    )
    adapter = HTTPAdapter(max_retries=retries, pool_connections=20, pool_maxsize=20)
    s.mount("https://", adapter)
    s.mount("http://", adapter)

    return s

def polite_pause():
    """Pausa entre peticiones para no saturar el servidor."""
    time.sleep(POLITE_SLEEP)

def safe_get(session: requests.Session, url: str, **kwargs) -> Optional[requests.Response]:
    """GET con timeout y captura de errores; devuelve Response o None."""
    try:
        resp = session.get(url, timeout=REQUEST_TIMEOUT, **kwargs)
        if not resp.ok:
            logging.warning("GET no OK %s -> %s", url, resp.status_code)
            return None
        return resp
    except requests.RequestException as exc:
        logging.warning("Error de red en GET %s: %s", url, exc)
        return None

# =========================
# Scraping helpers
# =========================

def urljoin(base: str, href: str) -> str:
    return urllib.parse.urljoin(base, href)

def find_sections(soup: BeautifulSoup) -> List[Tuple[str, BeautifulSoup]]:
    """
    Encuentra secciones: (título, tabla) en la página índice.
    Selector tolerante: busca h2 con <strong> y toma la tabla siguiente.
    """
    sections: List[Tuple[str, BeautifulSoup]] = []
    for h2 in soup.find_all("h2"):
        strong_tag = h2.find("strong")
        if not strong_tag:
            continue
        title = strong_tag.get_text(strip=True)
        next_table = h2.find_next("table")
        if next_table:
            sections.append((title, next_table))
    return sections

def clean_headers(table: BeautifulSoup) -> List[str]:
    """Lee <th> y rellena vacíos como Unnamed_i."""
    headers = []
    for i, th in enumerate(table.find_all("th")):
        text = (th.get_text() or "").strip()
        headers.append(text if text else f"Unnamed_{i}")
    return headers

def extract_detail_link(name_cell: BeautifulSoup) -> Optional[str]:
    """
    Intenta obtener un enlace de detalle asociado a la fila.
    1) <a> dentro del primer <td>
    2) Enlace en un sibling cercano (<h3> o similar)
    """
    if not name_cell:
        return None
    a0 = name_cell.find("a", href=True)
    if a0:
        return a0["href"]
    sib = name_cell.next_sibling
    if sib and getattr(sib, "find", None):
        a1 = sib.find("a", href=True)
        if a1:
            return a1["href"]
    return None

def extract_image_url(detail_html: str, base: str) -> Optional[str]:
    """Localiza un <a data-fancybox ... href='...'> y devuelve URL absoluta."""
    dsoup = BeautifulSoup(detail_html, "html.parser")
    a_img = dsoup.select_one("a[data-fancybox][href]")
    if not a_img:
        return None
    href = a_img.get("href")
    if not href:
        return None
    return urljoin(base, href)

def download_image(session: requests.Session, url: str, dest_path: Path) -> bool:
    """Descarga binaria con streaming si no existe. Devuelve True si queda en disco."""
    try:
        if dest_path.exists():
            return True
        resp = session.get(url, stream=True, timeout=REQUEST_TIMEOUT)
        if not resp.ok:
            logging.warning("No se pudo descargar imagen %s -> %s", url, resp.status_code)
            return False
        with dest_path.open("wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        return True
    except requests.RequestException as exc:
        logging.warning("Error descargando imagen %s: %s", url, exc)
        return False

def process_table(
    table: BeautifulSoup,
    session: requests.Session,
    images_dir: Path
) -> pd.DataFrame:
    """
    Procesa una tabla HTML y retorna un DataFrame con columnas:
    Adquirido, columnas originales, Detail Link, Image URL, Image Path, Imagen
    """
    headers = clean_headers(table)
    out_cols = ["Adquirido"] + headers + ["Detail Link", "Image URL", "Image Path", "Imagen"]
    rows = []

    for tr in table.find_all("tr"):
        tds = tr.find_all("td")
        if not tds:
            continue

        # Normaliza nº de celdas
        cells = [(td.get_text() or "").strip() for td in tds]
        if len(cells) < len(headers):
            cells += [""] * (len(headers) - len(cells))
        else:
            cells = cells[:len(headers)]

        row_data = ["No"] + cells

        name_cell = tds[0] if tds else None
        detail_link = extract_detail_link(name_cell)

        image_url = None
        image_path_str = None

        if detail_link:
            detail_url = urljoin(CHECKLIST_URL, detail_link)
            dresp = safe_get(session, detail_url)
            if dresp is not None:
                img_url = extract_image_url(dresp.text, SITE_BASE)
                if img_url:
                    image_url = img_url
                    image_name = Path(urllib.parse.urlparse(img_url).path).name
                    image_path = images_dir / image_name
                    if download_image(session, img_url, image_path):
                        image_path_str = str(image_path)
                polite_pause()

        row_data += [detail_link, image_url, image_path_str, ""]
        rows.append(row_data)

    df = pd.DataFrame(rows, columns=out_cols)

    # Elimina todas las columnas Unnamed_* si quedaron
    unnamed = [c for c in df.columns if c.startswith("Unnamed_")]
    if unnamed:
        df = df.drop(columns=unnamed)

    return df

# =========================
# Excel: lectura y fusión
# =========================

def sanitize_sheet_base(title: str) -> str:
    """
    Limpia el título para usar como base de nombre de hoja:
    - Sustituye caracteres inválidos.
    - Mantiene letras/números/espacios/guiones bajos.
    - Colapsa espacios a guiones bajos.
    - Recorta a 31 chars (luego se ajusta con sufijos si hay colisión).
    """
    # Reemplazos básicos
    t = title.replace("/", "-").replace("\\", "-").replace(":", "-")
    t = t.replace("?", "").replace("*", "").replace("[", "(").replace("]", ")")
    # Solo letras/números/espacios/guiones/guion_bajo
    t = re.sub(r"[^A-Za-z0-9 \-_]", "", t)
    # Colapsa espacios a "_"
    t = re.sub(r"\s+", "_", t.strip())
    if not t:
        t = "Sheet"
    return t[:31]  # base inicial (luego puede acortarse para sufijos)

def unique_sheet_name(title: str, used: Set[str]) -> str:
    """
    Genera un nombre de hoja único (case-insensitive) de <=31 caracteres.
    - Usa una base saneada del título.
    - Si existe, añade sufijos _2, _3, ... recortando base si es necesario.
    - Evita colisiones case-insensitive.
    - Como último recurso añade un hash corto.
    """
    base = sanitize_sheet_base(title)
    name = base
    norm = name.lower()
    if norm not in used:
        used.add(norm)
        return name

    # Añade sufijos incrementales
    counter = 2
    while True:
        suffix = f"_{counter}"
        maxlen = 31 - len(suffix)
        candidate = (base[:maxlen] if maxlen > 0 else base[:31]) + suffix
        norm_c = candidate.lower()
        if norm_c not in used:
            used.add(norm_c)
            return candidate
        counter += 1
        # Seguridad: si crecemos demasiado, añadimos hash
        if counter > 99:
            h = hashlib.sha1(title.encode("utf-8")).hexdigest()[:4]
            base2 = (base[:31 - 5]) + "_" + h  # deja espacio para _2/_3
            base = base2
            counter = 2  # reintenta con base hasheada

def read_existing_excel(file_path: Path) -> List[Tuple[str, pd.DataFrame]]:
    """
    Lee el Excel existente, cada hoja con:
      A1 => título completo,
      fila 2 => cabeceras,
      filas >=3 => datos.
    """
    if not file_path.exists():
        logging.info("No existe Excel previo, se creará uno nuevo.")
        return []

    wb = load_workbook(file_path)
    sections_data_old: List[Tuple[str, pd.DataFrame]] = []

    for ws in wb.worksheets:
        title = ws.cell(row=1, column=1).value
        if not title:
            continue

        headers = []
        col = 1
        while True:
            val = ws.cell(row=2, column=col).value
            if val is None:
                break
            headers.append(val)
            col += 1

        data_rows = []
        r = 3
        max_row = ws.max_row
        while r <= max_row:
            row_values = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
            if all(x is None for x in row_values):
                break
            data_rows.append(row_values)
            r += 1

        df_section = pd.DataFrame(data_rows, columns=headers)
        sections_data_old.append((title, df_section))

    logging.info("Excel previo leído con éxito.")
    return sections_data_old

def make_key(df: pd.DataFrame) -> pd.DataFrame:
    """
    Crea una clave compuesta estable para el merge.
    Usa columnas disponibles entre ["Name", "Wave", "Year"].
    Versión robusta: usa .apply(..., axis=1) para garantizar una Serie.
    """
    df = df.copy()
    candidates = [c for c in ["Name", "Wave", "Year"] if c in df.columns]
    if not candidates:
        df["__key__"] = df.index.astype(str)
        return df

    def _join_row(row):
        return "|".join([str(row[c]).strip() if pd.notna(row[c]) else "" for c in candidates])

    df["__key__"] = df.apply(_join_row, axis=1)
    return df

def combine_sections(
    sections_new: List[Tuple[str, pd.DataFrame]],
    sections_old: List[Tuple[str, pd.DataFrame]]
) -> List[Tuple[str, pd.DataFrame]]:
    """
    Combina datos nuevos y antiguos por sección.
    - Alinea columnas.
    - Merge por clave compuesta (__key__) preservando orden web.
    - Mantiene filas antiguas no presentes en la web.
    - Respeta "Adquirido = Sí" (prioriza fila antigua completa).
    """
    old_map = {t: df for t, df in sections_old}
    result: List[Tuple[str, pd.DataFrame]] = []

    def _dedupe_index(dfk: pd.DataFrame) -> pd.DataFrame:
        """Desambiguación de claves duplicadas en el índice."""
        if not dfk.index.has_duplicates:
            return dfk
        counts = {}
        new_index = []
        for k in dfk.index:
            n = counts.get(k, 0)
            new_index.append(k if n == 0 else f"{k}#{n}")
            counts[k] = n + 1
        dfk = dfk.copy()
        dfk.index = new_index
        return dfk

    for title, df_new in sections_new:
        df_old = old_map.get(title, pd.DataFrame())

        # Alinea columnas (mantener todas)
        new_cols = df_new.columns.tolist()
        old_cols = df_old.columns.tolist()
        missing_in_new = [c for c in old_cols if c not in new_cols]
        final_cols = new_cols + missing_in_new

        df_new = df_new.reindex(columns=final_cols, fill_value="")
        df_old = df_old.reindex(columns=final_cols, fill_value="")

        # Merge por clave compuesta
        df_new_k = make_key(df_new).set_index("__key__", drop=False)
        df_old_k = make_key(df_old).set_index("__key__", drop=False)

        # Desambiguar si hubiera claves duplicadas
        df_new_k = _dedupe_index(df_new_k)
        df_old_k = _dedupe_index(df_old_k)

        final_rows = []

        # Primero, filas en orden web (nuevas)
        for k in df_new_k.index:
            if k in df_old_k.index:
                old_row = df_old_k.loc[k].copy()
                if str(old_row.get("Adquirido", "No")) == "Sí":
                    final_rows.append(old_row)
                else:
                    new_row = df_new_k.loc[k].copy()
                    new_row["Adquirido"] = old_row.get("Adquirido", "No")
                    final_rows.append(new_row)
            else:
                new_row = df_new_k.loc[k].copy()
                new_row["Adquirido"] = "No"
                final_rows.append(new_row)

        # Luego, filas antiguas que ya no aparecen en la web
        for k in df_old_k.index:
            if k not in df_new_k.index:
                final_rows.append(df_old_k.loc[k].copy())

        df_final = pd.DataFrame(final_rows, columns=final_cols)

        # Reordenar columnas según DESIRED_ORDER
        existing_in_desired = [c for c in DESIRED_ORDER if c in df_final.columns]
        remaining_cols = [c for c in df_final.columns if c not in existing_in_desired]
        df_final = df_final.reindex(columns=existing_in_desired + remaining_cols)

        result.append((title, df_final))

    return result

# =========================
# Excel: escritura (una pasada, con hipervínculos)
# =========================

def write_excel_with_links(
    excel_path: Path,
    sections: List[Tuple[str, pd.DataFrame]]
) -> None:
    """
    Escribe cada sección en una hoja, con título en A1 y validación/formatos.
    Inserta hipervínculos a las imágenes (si existe el fichero en disco).
    Todo en una sola pasada. Garantiza nombres de hoja únicos.
    """
    with pd.ExcelWriter(excel_path, engine="xlsxwriter") as writer:
        wb = writer.book
        title_fmt = wb.add_format({"bg_color": "#ADD8E6", "bold": True})
        fmt_green = wb.add_format({"bg_color": "#C6EFCE", "font_color": "#006100"})
        fmt_red = wb.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006"})

        used_names: Set[str] = set()  # nombres (lowercase) ya usados

        for title, df in sections:
            sheet_name = unique_sheet_name(title, used_names)
            ws = wb.add_worksheet(sheet_name)
            writer.sheets[sheet_name] = ws  # evita colisión con to_excel

            # Título en A1 (fila 0)
            num_cols = max(1, df.shape[1])
            ws.merge_range(0, 0, 0, num_cols - 1, title, title_fmt)

            # DataFrame desde fila 2 (índice 1)
            df.to_excel(writer, sheet_name=sheet_name, startrow=1, startcol=0, index=False)

            df_rows = df.shape[0]
            if df_rows == 0:
                continue

            # Validación/formatos en "Adquirido"
            if "Adquirido" in df.columns:
                col_idx = df.columns.get_loc("Adquirido")
                data_start = 2
                data_end = 1 + df_rows

                ws.data_validation(
                    data_start, col_idx, data_end, col_idx,
                    {"validate": "list", "source": ["Sí", "No"]}
                )

                col_letter = xlsxwriter.utility.xl_col_to_name(col_idx)
                data_range = f"{col_letter}{data_start+1}:{col_letter}{data_end+1}"

                ws.conditional_format(data_range, {
                    "type": "cell", "criteria": "==", "value": '"Sí"', "format": fmt_green
                })
                ws.conditional_format(data_range, {
                    "type": "cell", "criteria": "==", "value": '"No"', "format": fmt_red
                })

            # Hipervínculos a imágenes
            if "Imagen" in df.columns and "Image Path" in df.columns:
                imagen_col = df.columns.get_loc("Imagen")
                path_col = df.columns.get_loc("Image Path")
                data_start = 2
                for i in range(df_rows):
                    img_path = df.iat[i, path_col]
                    if img_path:
                        p = Path(img_path)
                        if p.exists():
                            link = "file:///" + p.as_posix()
                            ws.write_url(data_start + i, imagen_col, link, string="Ver Imagen")

# =========================
# Pipeline principal
# =========================

def main() -> None:
    start_time = time.time()

    project_root, excel_path, images_dir = get_project_paths()
    logging.info("Raíz del proyecto: %s", project_root)
    logging.info("Excel de salida:   %s", excel_path)
    logging.info("Carpeta imágenes:  %s", images_dir)

    session = build_session()

    logging.info("Descargando página índice...")
    resp = safe_get(session, CHECKLIST_URL)
    if resp is None:
        logging.error("No se pudo acceder a la página índice.")
        return
    soup = BeautifulSoup(resp.text, "html.parser")

    logging.info("Buscando secciones...")
    sections_html = find_sections(soup)
    logging.info("Secciones encontradas: %d", len(sections_html))

    # Procesar tablas
    sections_new: List[Tuple[str, pd.DataFrame]] = []
    for title, tbl in sections_html:
        logging.info("Procesando sección: %s", title)
        df = process_table(tbl, session, images_dir)
        sections_new.append((title, df))
        polite_pause()

    # Leer Excel previo (si existe)
    sections_old = read_existing_excel(excel_path)

    # Combinar
    sections_final = combine_sections(sections_new, sections_old)

    # Escribir Excel con hipervínculos (una sola pasada) con nombres únicos
    write_excel_with_links(excel_path, sections_final)

    elapsed = time.time() - start_time
    logging.info("Extracción completada en %s", time.strftime("%H:%M:%S", time.gmtime(elapsed)))
    logging.info("Resultado guardado en: %s", excel_path)

def get_scraped_data() -> List[Tuple[str, pd.DataFrame]]:
    """
    Función expuesta para la Clean Architecture.
    Devuelve los datos scrapeados sin interactuar con el Excel.
    """
    project_root, excel_path, images_dir = get_project_paths()
    session = build_session()
    
    logging.info("Clean Arch: Descargando Checklist...")
    resp = safe_get(session, CHECKLIST_URL)
    if resp is None:
        return []
        
    soup = BeautifulSoup(resp.text, "html.parser")
    sections_html = find_sections(soup)
    
    sections_data = []
    for title, tbl in sections_html:
        # Para la importación no necesitamos descargar todas las imágenes obligatoriamente,
        # pero mantenemos la lógica si ya está ahí.
        df = process_table(tbl, session, images_dir)
        sections_data.append((title, df))
        polite_pause()
        
    return sections_data


if __name__ == "__main__":
    main()